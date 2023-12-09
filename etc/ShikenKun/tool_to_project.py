#!python3
# coding: utf-8


import shutil
import openpyxl
import json


with open('settings.json', 'r') as fp:
    jData = json.load(fp)
    src_book                = jData['src']['book']
    src_row                 = jData['src']['row']
    src_premise             = jData['src']['premise']
    src_procedure           = jData['src']['procedure']
    src_confirmation        = jData['src']['confirmation']
    filter_level            = jData['src']['filter_level']
    filter_xxxxx            = jData['src']['filter_xxxxx']
    dst_book                = jData['dst']['book']
    dst_row                 = jData['dst']['row']
    dst_premise             = jData['dst']['premise']
    dst_procedure           = jData['dst']['procedure']
    dst_confirmation        = jData['dst']['confirmation']


def main():
    src = openpyxl.load_workbook('./source_test_data.xlsx')
    dst = openpyxl.load_workbook('./template.xlsx')

    srcSheet = src[src.sheetnames[src_book['sheet']]]
    dstSheet = dst[dst.sheetnames[dst_book['sheet']]]

    dstColumn = dst_row['start']

    for row in range(src_row['start'], srcSheet.max_row):

        if (filter_level['enable']):
            if (srcSheet.cell(row=row, column=filter_level['column']).value >= filter_level['input']):
                continue
        
        if (filter_xxxxx['enable']):
            if (srcSheet.cell(row=row, column=filter_xxxxx['column']).value >= filter_xxxxx['input']):
                continue

        value = ''
        for column in range(src_premise['start'], src_premise['end']):
            cell = srcSheet.cell(row=row, column=column)
            if (cell.value != None):
                value += '・'
                value += cell.value
                value += '\r\n'
        dstSheet.cell(row=dstColumn, column=dst_premise).value = value
        dstSheet.cell(row=dstColumn, column=dst_premise).alignment = openpyxl.styles.Alignment(wrapText=True)

        value = ''
        num = 1
        for column in range(src_procedure['start'], src_procedure['end']):
            cell = srcSheet.cell(row=row, column=column)
            if (cell.value != None):
                value += str(num)
                value += '.'
                value += cell.value
                value += '\r\n'
                num += 1
        dstSheet.cell(row=dstColumn, column=dst_procedure).value = value
        dstSheet.cell(row=dstColumn, column=dst_procedure).alignment = openpyxl.styles.Alignment(wrapText=True)
        
        value = ''
        for column in range(src_confirmation['start'], src_confirmation['end']):
            cell = srcSheet.cell(row=row, column=column)
            if (cell.value != None):
                value += '・'
                value += cell.value
                value += '\r\n'
        dstSheet.cell(row=dstColumn, column=dst_confirmation).value = value
        dstSheet.cell(row=dstColumn, column=dst_confirmation).alignment = openpyxl.styles.Alignment(wrapText=True)
        
        dstColumn += 1
    dst.save('dust_test_data.xlsx')


if __name__ == '__main__':
    print('[S]tool_to_project')
    main()
    print('[E]tool_to_project')
